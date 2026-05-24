"""Exportar/importar 'escenarios' completos como JSON.

Un escenario contiene TODO lo configurable desde la Tabla de mando
(modelo de cada centro, socios, personal, gastos) + ajustes globales
(tasa de IS, clave de prorrateo HQ, aperturas).

NO contiene el libro diario real (ese se sube por separado desde el
boton 📥 Libro diario). Tampoco contiene el cash flow: este se deriva
automaticamente de la P&L cuando se importa un escenario.

Uso tipico: varios socios bajan el JSON del escenario base, cada uno
juega con su propia version en local/cloud, y comparten su JSON entre
ellos sin tocar la data principal.
"""

from __future__ import annotations

import datetime as _dt
import io
import json
from typing import Any, Dict

from sqlalchemy import text

from . import config, store

SCHEMA_VERSION = 1


def export_scenario() -> bytes:
    """Devuelve un JSON (bytes) con toda la configuracion editable."""
    eng = store.get_engine()
    out: Dict[str, Any] = {
        "_schema": SCHEMA_VERSION,
        "_exported_at": _dt.datetime.utcnow().isoformat() + "Z",
        "centers": {},
        "globals": {},
    }
    with eng.begin() as conn:
        cm_rows = conn.execute(text(
            "SELECT centro, clave, valor FROM center_model")).fetchall()
        sp_rows = conn.execute(text(
            "SELECT centro, periodo, altas, churn, bajas "
            "FROM socios_plan")).fetchall()
        pp_rows = conn.execute(text(
            "SELECT centro, rol, nombre, bruto, mes_inicio, destino, "
            "ss_pct FROM personal_plan")).fetchall()
        gp_rows = conn.execute(text(
            "SELECT centro, idx, partida, importe, intervalo, apartado, "
            "mes_inicio FROM gastos_plan ORDER BY centro, idx")).fetchall()
        c_rows = conn.execute(text(
            "SELECT centro, apertura FROM centers")).fetchall()
        m_rows = conn.execute(text(
            "SELECT clave, valor FROM meta WHERE clave IN "
            "('is_rate', 'proration_key')")).fetchall()

    # Agrupacion por centro
    for c, lbl, _ in [(c, config.CENTER_LABELS.get(c, c), None)
                      for c in config.ALL_CENTERS]:
        out["centers"][c] = {
            "label": lbl,
            "apertura": "",
            "model_keys": {},
            "socios_plan": {},
            "personal": [],
            "gastos": [],
        }

    for centro, clave, valor in cm_rows:
        if centro in out["centers"]:
            out["centers"][centro]["model_keys"][clave] = valor

    for centro, periodo, altas, churn, bajas in sp_rows:
        if centro in out["centers"]:
            out["centers"][centro]["socios_plan"][periodo] = {
                "altas": float(altas or 0),
                "churn": float(churn or 0),
                "bajas": float(bajas or 0),
            }

    for centro, rol, nombre, bruto, mi, dest, ssp in pp_rows:
        if centro in out["centers"]:
            out["centers"][centro]["personal"].append({
                "rol": rol, "nombre": nombre or "",
                "bruto": float(bruto or 0),
                "mes_inicio": int(mi or 1),
                "destino": dest or "",
                "ss_pct": float(ssp or 0),
            })

    for centro, idx, partida, imp, itv, ap, mi in gp_rows:
        if centro in out["centers"]:
            out["centers"][centro]["gastos"].append({
                "partida": partida or "",
                "importe": float(imp or 0),
                "intervalo": int(itv or 1),
                "apartado": ap or "",
                "mes_inicio": int(mi or 1),
            })

    for centro, apertura in c_rows:
        if centro in out["centers"]:
            out["centers"][centro]["apertura"] = apertura or ""

    for clave, valor in m_rows:
        try:
            out["globals"][clave] = json.loads(valor)
        except (ValueError, TypeError):
            out["globals"][clave] = valor

    return json.dumps(out, indent=2, ensure_ascii=False).encode("utf-8")


def import_scenario(data: bytes) -> dict:
    """Aplica un escenario a la base. Devuelve un dict con un resumen
    de lo importado. Lanza ValueError si el JSON es invalido."""
    try:
        scenario = json.loads(data.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as e:
        raise ValueError("JSON invalido: %s" % e)

    if not isinstance(scenario, dict) or "centers" not in scenario:
        raise ValueError("El JSON no es un escenario valido "
                         "(falta la clave 'centers').")
    if scenario.get("_schema") != SCHEMA_VERSION:
        raise ValueError("Version de esquema incompatible "
                         "(esperado %d, recibido %r). Genera un nuevo "
                         "export desde la version actual de la app."
                         % (SCHEMA_VERSION, scenario.get("_schema")))

    eng = store.get_engine()
    summary = {"centros": 0, "gastos": 0, "personal": 0,
               "socios": 0, "aperturas": 0, "model_keys": 0}

    with eng.begin() as conn:
        # 1) center_model: borrar y reinsertar por cada centro presente
        for centro, data_c in scenario["centers"].items():
            if centro not in config.CENTERS and centro != config.HQ:
                continue
            summary["centros"] += 1
            conn.execute(
                text("DELETE FROM center_model WHERE centro=:c"),
                {"c": centro})
            model_keys = data_c.get("model_keys", {})
            for k, v in model_keys.items():
                conn.execute(text(
                    "INSERT INTO center_model (centro, clave, valor) "
                    "VALUES (:c, :k, :v)"),
                    {"c": centro, "k": str(k), "v": float(v)})
                summary["model_keys"] += 1

            # 2) socios_plan
            conn.execute(
                text("DELETE FROM socios_plan WHERE centro=:c"),
                {"c": centro})
            for periodo, plan in data_c.get("socios_plan", {}).items():
                conn.execute(text(
                    "INSERT INTO socios_plan "
                    "(centro, periodo, altas, churn, bajas) "
                    "VALUES (:c, :p, :a, :ch, :b)"),
                    {"c": centro, "p": str(periodo),
                     "a": float(plan.get("altas", 0) or 0),
                     "ch": float(plan.get("churn", 0) or 0),
                     "b": float(plan.get("bajas", 0) or 0)})
                summary["socios"] += 1

            # 3) personal
            conn.execute(
                text("DELETE FROM personal_plan WHERE centro=:c"),
                {"c": centro})
            for p in data_c.get("personal", []):
                conn.execute(text(
                    "INSERT INTO personal_plan "
                    "(centro, rol, nombre, bruto, mes_inicio, "
                    "destino, ss_pct) "
                    "VALUES (:c, :r, :n, :b, :m, :d, :s)"),
                    {"c": centro, "r": str(p.get("rol", "")),
                     "n": str(p.get("nombre", "") or ""),
                     "b": float(p.get("bruto", 0) or 0),
                     "m": int(p.get("mes_inicio", 1) or 1),
                     "d": str(p.get("destino", "") or ""),
                     "s": float(p.get("ss_pct", 0) or 0)})
                summary["personal"] += 1

            # 4) gastos
            conn.execute(
                text("DELETE FROM gastos_plan WHERE centro=:c"),
                {"c": centro})
            for i, g in enumerate(data_c.get("gastos", [])):
                conn.execute(text(
                    "INSERT INTO gastos_plan "
                    "(centro, idx, partida, importe, intervalo, "
                    "apartado, mes_inicio) "
                    "VALUES (:c, :i, :p, :imp, :it, :ap, :mi)"),
                    {"c": centro, "i": i,
                     "p": str(g.get("partida", "") or ""),
                     "imp": float(g.get("importe", 0) or 0),
                     "it": int(g.get("intervalo", 1) or 1),
                     "ap": str(g.get("apartado", "") or ""),
                     "mi": int(g.get("mes_inicio", 1) or 1)})
                summary["gastos"] += 1

            # 5) apertura (si viene)
            apertura = data_c.get("apertura")
            if apertura:
                conn.execute(text(
                    "UPDATE centers SET apertura=:a WHERE centro=:c"),
                    {"a": str(apertura), "c": centro})
                summary["aperturas"] += 1

        # 6) globals (is_rate, proration_key)
        for k, v in (scenario.get("globals") or {}).items():
            if k not in ("is_rate", "proration_key"):
                continue
            conn.execute(text("DELETE FROM meta WHERE clave=:k"),
                         {"k": k})
            conn.execute(text(
                "INSERT INTO meta (clave, valor) VALUES (:k, :v)"),
                {"k": k, "v": json.dumps(v)})

    return summary


def suggest_filename() -> str:
    ts = _dt.datetime.utcnow().strftime("%Y%m%d_%H%M")
    return "Kratos_escenario_%s.json" % ts


# --- P&L: exportar la P&L vista + importar overrides directos -------------
PNL_SCHEMA = 1


def export_pnl(model) -> bytes:
    """Exporta la P&L (todos los meses, por centro/apartado/partida) en
    JSON. El cliente edita los meses futuros y reimporta como overrides.
    `model` = pipeline.Model ya construido."""
    out: Dict[str, Any] = {
        "_schema": PNL_SCHEMA,
        "_type": "pnl",
        "_exported_at": _dt.datetime.utcnow().isoformat() + "Z",
        "last_actual_period": model.last_actual_period,
        "nota": ("Edita SOLO los meses posteriores a last_actual_period. "
                 "Los meses <= last_actual son reales del libro diario y "
                 "se ignoran al reimportar. Los importes de gasto van en "
                 "negativo, los ingresos en positivo."),
        "centros": {},
    }
    df = model.pnl_long
    if df is None or df.empty:
        return json.dumps(out, indent=2, ensure_ascii=False).encode("utf-8")

    for centro in config.ALL_CENTERS:
        sub = df[df["centro"] == centro]
        if sub.empty:
            continue
        lineas = []
        grp = sub.groupby(["apartado", "partida"])
        for (apartado, partida), g in grp:
            valores = {p: round(float(v), 2)
                       for p, v in zip(g["periodo"], g["valor"])}
            lineas.append({
                "apartado": apartado, "partida": partida,
                "valores": valores})
        out["centros"][centro] = {
            "label": config.CENTER_LABELS.get(centro, centro),
            "lineas": lineas,
        }
    return json.dumps(out, indent=2, ensure_ascii=False).encode("utf-8")


def import_pnl_overrides(data: bytes, last_actual: str) -> dict:
    """Lee un JSON de P&L (formato export_pnl, editado) y guarda los
    valores de los meses FUTUROS como overrides. Devuelve resumen."""
    try:
        doc = json.loads(data.decode("utf-8"))
    except (UnicodeDecodeError, json.JSONDecodeError) as e:
        raise ValueError("JSON invalido: %s" % e)
    if not isinstance(doc, dict) or doc.get("_type") != "pnl":
        raise ValueError("El JSON no es una P&L exportada desde la app "
                         "(falta _type='pnl').")
    if doc.get("_schema") != PNL_SCHEMA:
        raise ValueError("Version de esquema incompatible. Re-exporta la "
                         "P&L desde la version actual de la app.")

    rows = []
    last_actual = last_actual or ""
    for centro, cdata in (doc.get("centros") or {}).items():
        if centro not in config.ALL_CENTERS:
            continue
        for ln in cdata.get("lineas", []):
            apartado = str(ln.get("apartado", ""))
            partida = str(ln.get("partida", ""))
            for periodo, valor in (ln.get("valores") or {}).items():
                if str(periodo) <= last_actual:
                    continue          # nunca pisa lo real
                rows.append({
                    "centro": centro, "periodo": str(periodo),
                    "apartado": apartado, "partida": partida,
                    "valor": float(valor or 0)})

    store.set_pnl_overrides(rows, replace=True)
    centros = sorted({r["centro"] for r in rows})
    return {"overrides": len(rows), "centros": len(centros),
            "centros_list": centros}


def suggest_pnl_filename() -> str:
    ts = _dt.datetime.utcnow().strftime("%Y%m%d_%H%M")
    return "Kratos_PL_%s.json" % ts


# --- P&L en EXCEL (lo que el cliente edita en Excel) ----------------------
_APARTADO_ORDER = [
    "Ingresos",
    "Aprovisionamientos",
    "Gastos de Personal",
    "Otros gastos de explotacion",
    "Gastos de estructura HQ",
    "Amortizaciones",
    "Gastos financieros",
]


def _apartado_rank(ap: str) -> int:
    try:
        return _APARTADO_ORDER.index(ap)
    except ValueError:
        return len(_APARTADO_ORDER)


def export_pnl_xlsx(model) -> bytes:
    """P&L en Excel: una hoja por centro, filas = (apartado, partida),
    columnas = meses. Los meses REALES (<= last_actual) van sombreados en
    amarillo (no editar); los futuros en blanco (editables). Al reimportar
    solo se aplican los meses futuros."""
    import xlsxwriter
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {"in_memory": True})
    f_title = wb.add_format({"bold": True, "font_size": 13})
    f_note = wb.add_format({"italic": True, "font_color": "#555555",
                            "text_wrap": True})
    f_hdr = wb.add_format({"bold": True, "bg_color": "#1F4E78",
                           "font_color": "white", "border": 1,
                           "align": "center"})
    f_key = wb.add_format({"border": 1, "bg_color": "#EEEEEE"})
    f_lbl = wb.add_format({"border": 1})
    f_real = wb.add_format({"border": 1, "bg_color": "#FFF2CC",
                            "num_format": "#,##0"})       # real - no editar
    f_fut = wb.add_format({"border": 1, "num_format": "#,##0"})   # editable

    df = model.pnl_long
    last_actual = model.last_actual_period or ""

    for centro in config.ALL_CENTERS:
        ws = wb.add_worksheet(centro)
        ws.set_column(0, 0, 22)
        ws.set_column(1, 1, 30)
        ws.write(0, 1, "P&L — %s" %
                 config.CENTER_LABELS.get(centro, centro), f_title)
        ws.write(1, 1,
                 "Celdas AMARILLAS = meses reales del libro diario "
                 "(no editar, se ignoran al reimportar). Celdas BLANCAS "
                 "= meses futuros: edítalos y reimporta. Gastos en "
                 "negativo, ingresos en positivo. No cambies las columnas "
                 "A (apartado) ni B (partida).", f_note)
        ws.set_row(1, 44)

        periods = model.center_periods.get(centro, model.periods)
        ws.set_column(2, 2 + len(periods), 11)

        r = 3
        ws.write(r, 0, "apartado", f_hdr)
        ws.write(r, 1, "partida", f_hdr)
        for j, p in enumerate(periods):
            ws.write(r, 2 + j, p, f_hdr)
        ws.freeze_panes(r + 1, 2)
        r += 1

        sub = df[df["centro"] == centro] if df is not None else None
        if sub is None or sub.empty:
            continue
        # valores por (apartado, partida, periodo)
        vmap = {}
        for ap, part, per, val in zip(sub["apartado"], sub["partida"],
                                      sub["periodo"], sub["valor"]):
            vmap[(ap, part, per)] = float(val)
        combos = sorted(
            {(ap, part) for ap, part in zip(sub["apartado"],
                                            sub["partida"])},
            key=lambda x: (_apartado_rank(x[0]), x[1]))
        for ap, part in combos:
            ws.write(r, 0, ap, f_key)
            ws.write(r, 1, part, f_lbl)
            for j, p in enumerate(periods):
                v = vmap.get((ap, part, p), 0.0)
                fmt = f_real if p <= last_actual else f_fut
                ws.write_number(r, 2 + j, round(v, 2), fmt)
            r += 1

    wb.close()
    buf.seek(0)
    return buf.read()


def import_pnl_xlsx(data: bytes, last_actual: str) -> dict:
    """Lee el Excel de P&L (formato export_pnl_xlsx, editado) y guarda los
    valores de los meses FUTUROS como overrides."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    last_actual = last_actual or ""
    rows = []
    for centro in config.ALL_CENTERS:
        if centro not in wb.sheetnames:
            continue
        ws = wb[centro]
        # Fila de cabecera: la que tiene 'apartado' en col A
        header_row = None
        for ridx in range(1, min(ws.max_row, 12) + 1):
            if str(ws.cell(ridx, 1).value or "").strip().lower() == \
                    "apartado":
                header_row = ridx
                break
        if header_row is None:
            continue
        # Mapear columnas de mes (desde col 3)
        periods = {}
        c = 3
        while True:
            val = ws.cell(header_row, c).value
            if val is None or str(val).strip() == "":
                break
            periods[c] = str(val).strip()
            c += 1
        # Filas de datos
        for ridx in range(header_row + 1, ws.max_row + 1):
            ap = ws.cell(ridx, 1).value
            part = ws.cell(ridx, 2).value
            if not ap or not part:
                continue
            for col, per in periods.items():
                if per <= last_actual:
                    continue
                cell = ws.cell(ridx, col).value
                if cell is None or cell == "":
                    continue
                try:
                    valor = float(cell)
                except (TypeError, ValueError):
                    continue
                rows.append({
                    "centro": centro, "periodo": per,
                    "apartado": str(ap).strip(),
                    "partida": str(part).strip(), "valor": valor})

    store.set_pnl_overrides(rows, replace=True)
    centros = sorted({r["centro"] for r in rows})
    return {"overrides": len(rows), "centros": len(centros),
            "centros_list": centros}


def suggest_pnl_xlsx_filename() -> str:
    ts = _dt.datetime.utcnow().strftime("%Y%m%d_%H%M")
    return "Kratos_PL_%s.xlsx" % ts
