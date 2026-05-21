"""Lee y normaliza el libro diario de Holded (XLSX).

Robusto a:
- filas de metadatos antes de la cabecera (deteccion dinamica),
- fila final "Informe creado automaticamente con Holded ...",
- fechas como datetime o como texto dd/mm/aaaa,
- importes como numero o como texto con coma decimal,
- nombres de cuenta con espacios/acentos irregulares,
- tags multiples con ruido de gestoria.
"""

from __future__ import annotations

import datetime as _dt
import unicodedata
from dataclasses import dataclass
from typing import List, Optional

import openpyxl
import pandas as pd

from . import config

# Cabecera de Holded -> nombre canonico interno
_HEADER_MAP = {
    "asiento": "asiento",
    "linea": "linea",
    "fecha": "fecha",
    "tipo": "tipo",
    "descripcion": "descripcion",
    "documento": "documento",
    "cuenta": "cuenta",
    "nombre de la cuenta": "cuenta_nombre",
    "debe": "debe",
    "haber": "haber",
    "tags": "tags",
    "punteado": "punteado",
}

_ACCENTS = str.maketrans("ÁÀÄÂÃÉÈËÊÍÌÏÎÓÒÖÔÕÚÙÜÛÑÇ", "AAAAAEEEEIIIIOOOOOUUUUNC")


def norm_text(s: object) -> str:
    """Mayusculas, sin acentos, espacios colapsados (para emparejar nombres)."""
    if s is None:
        return ""
    t = str(s).strip()
    t = unicodedata.normalize("NFKC", t)
    t = " ".join(t.split())
    return t.upper().translate(_ACCENTS)


def _norm_header(s: object) -> str:
    return norm_text(s).lower()


def to_float(v: object) -> float:
    """Convierte numero o texto ('1.234,56', '1,234.56', '12,5 €') a float."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        try:
            f = float(v)
        except (TypeError, ValueError):
            return 0.0
        return 0.0 if pd.isna(f) else f
    t = str(v).strip().replace(" ", " ")
    for ch in ("€", "$", " "):
        t = t.replace(ch, "")
    if not t:
        return 0.0
    has_comma, has_dot = "," in t, "." in t
    if has_comma and has_dot:
        # el ultimo separador es el decimal
        if t.rfind(",") > t.rfind("."):
            t = t.replace(".", "").replace(",", ".")
        else:
            t = t.replace(",", "")
    elif has_comma:
        t = t.replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return 0.0


def _to_period(v: object) -> Optional[str]:
    if isinstance(v, (_dt.datetime, _dt.date)):
        return "%04d-%02d" % (v.year, v.month)
    if v is None:
        return None
    t = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            d = _dt.datetime.strptime(t[:10], fmt)
            return "%04d-%02d" % (d.year, d.month)
        except ValueError:
            continue
    return None


def _to_date(v: object) -> Optional[_dt.date]:
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    if v is None:
        return None
    t = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
        try:
            return _dt.datetime.strptime(t[:10], fmt).date()
        except ValueError:
            continue
    return None


def _s(v: object) -> str:
    return "" if v is None else str(v).strip()


def _account_code(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, float):
        v = int(v)
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.zfill(8) if s.isdigit() and len(s) < 8 else s


def clean_tags(v: object) -> List[str]:
    """Lista de tags utiles (minusculas, sin ruido, sin duplicar, en orden)."""
    if v is None:
        return []
    out: List[str] = []
    for part in str(v).split(","):
        t = part.strip().lower()
        if t and t not in config.NOISE_TAGS and t not in out:
            out.append(t)
    return out


@dataclass
class IngestResult:
    df: pd.DataFrame
    n_rows: int
    sum_debe: float
    sum_haber: float
    date_min: Optional[_dt.date]
    date_max: Optional[_dt.date]
    sheet: str
    header_row: int
    dropped_rows: int


def _find_header(rows: List[tuple]) -> int:
    for i, row in enumerate(rows[:20]):
        cells = {_norm_header(c) for c in row if c is not None}
        if config.HEADER_REQUIRED_FIELDS.issubset(cells):
            return i
    raise ValueError(
        "No se encontro la fila de cabecera (Asiento/Fecha/Cuenta/Debe/Haber). "
        "Revisa que el archivo sea el libro diario exportado de Holded."
    )


def load_ledger(path: str, source_name: Optional[str] = None) -> IngestResult:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in config.SHEET_CANDIDATES if s in wb.sheetnames),
                 wb.sheetnames[0])
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    hidx = _find_header(rows)
    header = [_norm_header(c) for c in rows[hidx]]
    colmap = {pos: _HEADER_MAP[name] for pos, name in enumerate(header)
              if name in _HEADER_MAP}

    records = []
    dropped = 0
    for row in rows[hidx + 1:]:
        rec = {canon: row[pos] if pos < len(row) else None
               for pos, canon in colmap.items()}
        asi = rec.get("asiento")
        # fila valida = asiento numerico y cuenta presente
        if not isinstance(asi, (int, float)) or rec.get("cuenta") in (None, ""):
            dropped += 1
            continue
        debe = to_float(rec.get("debe"))
        haber = to_float(rec.get("haber"))
        cuenta = _account_code(rec.get("cuenta"))
        nombre = _s(rec.get("cuenta_nombre"))
        records.append({
            "asiento": int(asi),
            "linea": int(rec["linea"]) if isinstance(rec.get("linea"), (int, float)) else 0,
            "fecha": _to_date(rec.get("fecha")),
            "periodo": _to_period(rec.get("fecha")),
            "tipo": _s(rec.get("tipo")),
            "descripcion": _s(rec.get("descripcion")),
            "documento": _s(rec.get("documento")),
            "cuenta": cuenta,
            "prefijo2": cuenta[:2],
            "prefijo3": cuenta[:3],
            "cuenta_nombre": nombre,
            "nombre_norm": norm_text(nombre),
            "debe": debe,
            "haber": haber,
            "importe": round(debe - haber, 2),
            "tags": clean_tags(rec.get("tags")),
            "punteado": _s(rec.get("punteado")),
            "source_file": source_name or path.rsplit("/", 1)[-1],
        })

    df = pd.DataFrame.from_records(records)
    if df.empty:
        raise ValueError("El archivo no contiene asientos validos.")

    dates = [d for d in df["fecha"] if d is not None]
    return IngestResult(
        df=df,
        n_rows=len(df),
        sum_debe=round(float(df["debe"].sum()), 2),
        sum_haber=round(float(df["haber"].sum()), 2),
        date_min=min(dates) if dates else None,
        date_max=max(dates) if dates else None,
        sheet=sheet,
        header_row=hidx + 1,
        dropped_rows=dropped,
    )
