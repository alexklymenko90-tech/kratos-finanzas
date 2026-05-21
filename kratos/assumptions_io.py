"""Excel de supuestos que edita el cliente: exportar plantilla / importar.

Una pestana por centro (K1..K4, HQ). Filas = partidas (las mismas que la
P&L) + ajustes (CAPEX, anos amortizacion, peso % HQ, desfases). Columnas =
36 meses. La columna A lleva una CLAVE interna para que la importacion sea
robusta aunque el cliente cambie los textos.
"""

from __future__ import annotations

import io
from typing import Dict, List, Optional

import openpyxl
import pandas as pd
import xlsxwriter

from . import config, projections
from .pnl import month_range


def _key_partida(p: str) -> str:
    return "P|" + p


def _key_special(s: str) -> str:
    return "S|" + s


def export_template(opening: Dict[str, str],
                    start_month: str = config.DEFAULT_START_MONTH,
                    horizon: int = config.HORIZON_MONTHS,
                    last_actual_period: Optional[str] = None,
                    existing: Optional[pd.DataFrame] = None) -> bytes:
    periods = month_range(start_month, horizon)
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {"in_memory": True})

    f_title = wb.add_format({"bold": True, "font_size": 14})
    f_note = wb.add_format({"italic": True, "font_color": "#555555",
                            "text_wrap": True})
    f_hdr = wb.add_format({"bold": True, "bg_color": "#1F4E78",
                           "font_color": "white", "border": 1,
                           "align": "center"})
    f_sec = wb.add_format({"bold": True, "bg_color": "#D6E4F0", "border": 1})
    f_lbl = wb.add_format({"border": 1})
    f_num = wb.add_format({"border": 1, "num_format": "#,##0.00"})
    f_pre = wb.add_format({"border": 1, "bg_color": "#9E9E9E"})       # antes de apertura
    f_real = wb.add_format({"border": 1, "bg_color": "#FFF2CC",
                            "num_format": "#,##0.00"})                # mes ya real
    f_key = wb.add_format({"font_color": "#BBBBBB", "font_size": 8})

    exist = {}
    if existing is not None and not existing.empty:
        for _, r in existing.iterrows():
            exist[(r["centro"], r["concepto"], r["periodo"])] = r["valor"]

    inc = projections.income_partidas()
    exp = projections.expense_sections()

    for centro in config.CENTERS + [config.HQ]:
        ws = wb.add_worksheet(centro)
        ws.set_column(0, 0, 14, f_key)
        ws.set_column(1, 1, 34)
        ws.set_column(2, 2 + len(periods), 11)
        ap = opening.get(centro) or start_month
        ws.write(0, 1, "%s — Supuestos de proyeccion" %
                 config.CENTER_LABELS.get(centro, centro), f_title)
        ws.write(1, 1,
                 "Apertura: %s. Escribe el importe mensual ESPERADO. "
                 "Gastos en POSITIVO. Celdas grises = antes de apertura "
                 "(no escribir). Celdas amarillas = mes ya cerrado con "
                 "datos reales (lo que escribas ahi se ignora)." % ap,
                 f_note)
        ws.set_row(1, 46)

        r = 3
        ws.write(r, 0, "clave", f_key)
        ws.write(r, 1, "Concepto", f_hdr)
        for j, p in enumerate(periods):
            ws.write(r, 2 + j, p, f_hdr)
        r += 1

        def block(title: str, items: List[tuple]):
            nonlocal r
            ws.write(r, 1, title, f_sec)
            for j in range(len(periods)):
                ws.write_blank(r, 2 + j, None, f_sec)
            r += 1
            for key, label in items:
                ws.write(r, 0, key, f_key)
                ws.write(r, 1, label, f_lbl)
                for j, per in enumerate(periods):
                    fmt = f_num
                    if per < ap:
                        fmt = f_pre
                    elif last_actual_period and per <= last_actual_period:
                        fmt = f_real
                    val = exist.get((centro, key.split("|", 1)[1], per))
                    if val is not None and per >= ap:
                        ws.write_number(r, 2 + j, float(val), fmt)
                    else:
                        ws.write_blank(r, 2 + j, None, fmt)
                r += 1

        block("INGRESOS (importe mensual esperado, sin IVA)",
              [(_key_partida(p), p) for p in inc])
        for sec, parts in exp.items():
            block("GASTOS — %s (en positivo)" % sec,
                  [(_key_partida(p), p) for p in parts])
        block("INVERSION Y AJUSTES",
              [(_key_special(s), projections.SPECIAL_LABELS[s])
               for s in projections.SPECIALS])

        ws.freeze_panes(4, 2)

    wb.close()
    buf.seek(0)
    return buf.read()


def import_template(file_or_bytes) -> pd.DataFrame:
    """Devuelve long df: centro, periodo, concepto, valor."""
    if isinstance(file_or_bytes, (bytes, bytearray)):
        src = io.BytesIO(file_or_bytes)
    else:
        src = file_or_bytes
    wb = openpyxl.load_workbook(src, data_only=True)
    rows = []
    valid_centers = set(config.CENTERS + [config.HQ])
    for sheet in wb.sheetnames:
        if sheet not in valid_centers:
            continue
        ws = wb[sheet]
        grid = list(ws.iter_rows(values_only=True))
        # localizar fila de cabecera (col B == "Concepto")
        hidx = None
        for i, row in enumerate(grid[:10]):
            if row and str(row[1]).strip().lower() == "concepto":
                hidx = i
                break
        if hidx is None:
            continue
        periods = [str(c).strip() for c in grid[hidx][2:] if c]
        for row in grid[hidx + 1:]:
            if not row or not row[0]:
                continue
            key = str(row[0]).strip()
            if "|" not in key:
                continue
            kind, name = key.split("|", 1)
            for j, per in enumerate(periods):
                cell = row[2 + j] if (2 + j) < len(row) else None
                if cell is None or cell == "":
                    continue
                try:
                    val = float(cell)
                except (TypeError, ValueError):
                    continue
                if val == 0:
                    continue
                rows.append({"centro": sheet, "periodo": per,
                             "concepto": name, "valor": val})
    return pd.DataFrame(
        rows, columns=["centro", "periodo", "concepto", "valor"])
